import os
import fnmatch
from alive_progress import alive_bar
import time

import openpyxl as opx

# Путь к текущему рабочему каталогу
BASE_DIR = os.getcwd()
# Справочники
# ministries = {"m1": 'Министерство финансов РФ', "m2": 'Министерство обороны РФ', "m3": 'Министерство транспорта РФ'}
ministries = {"m1": 'Министерство финансов РФ'}
# year_dec = ('2012', '2013', '2014', '2015', '2016', '2017', '2018', '2019')
year_dec = ('2019',)
not_employees = ('Сотрудник', 'Супруг', 'Супруга', 'Несовершеннолетний ребенок')

property_dec = {"недвижимое в собственности": [3, 4, 5, 6], "недвижимое в пользовании": [7, 8, 9], "движимое": [10],
                "годовой доход": [11], "источник дохода": [12]}
own_dec = ('в собственности', 'в пользовании')


def isint(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def pars_xlsx():
    """
    Парсинг файлов формата xlsx
    :return:
    """
    global main_dec, main_dol, not_employees_dec, main_summ
    print('Парсинг файлов формата xlsx')

    for min_num, min_naim in ministries.items():
        """ министерство """
        # print(min_num)
        print(min_naim)
        for yd in year_dec:
            """ год """
            file_inp = f'input\\{min_num}_{yd}.xlsx'
            print(yd)

            # Предварительная подготовка
            # открываем файл с исходными данными
            wb = opx.load_workbook(os.path.join(BASE_DIR, file_inp))
            # выбор страницы если одна
            ws = wb.active
            # Строки, колонки
            i = 1

            print("> Поиск строки с первой записью")
            while i <= ws.max_row:
                """
                   Поиск строки с первой записью 
                """
                if ws.cell(i, 1).value == 1:
                    break
                i += 1

            all_dec = []  # список всей исходной информации

            print("> Создание списка сведений о декларациях")
            while i <= ws.max_row:
                """ Создание списка сведений о декларациях """
                num_col = 1
                row_dec = []  # список данных по строке
                while num_col <= ws.max_column:
                    # Заполним список
                    row_dec.append(ws.cell(i, num_col).value)
                    num_col += 1
                i += 1
                # проверка на отсутствие данных в строке
                if any(x is not None for x in row_dec[1:]):
                    all_dec.append(row_dec)

            print("> Заполнение новой таблицы")
            # Заполнение новой таблицы
            # открываем новый файл на запись
            wb_out = opx.load_workbook(os.path.join(BASE_DIR, "blank\\template.xlsx"))
            # выбор страницы
            ws_out = wb_out.active
            # Строки, столбецы
            num_row_out = 6
            num_col_out = 1
            num_row_inp = 1  # номер записи в исходном списке
            count = 0
            cn = 0  # номер записи в списке на выходе

            with alive_bar(len(all_dec)) as bar:
                while count < len(all_dec):
                    """ Преобразование предварительного списка """
                    # Заполнение списка в новом формате
                    if all_dec[count][0] == num_row_inp:
                        """ Главная запись по сотруднику """
                        # поиск записей по сотруднику
                        for row_range in range(len(all_dec[count:])):
                            """ Выделение диапазона строк по данному типу сотрудника """
                            if (all_dec[count + row_range][1] in not_employees[1:]) or \
                                    ((all_dec[count + row_range][0] != num_row_inp) and
                                     (all_dec[count + row_range][0] is not None)):
                                break

                        if (count + row_range) == (len(all_dec) - 1):
                            row_range += 1

                        row_dec_new = [cn, min_naim, yd, all_dec[count][1], all_dec[count][2]]
                        for prop in property_dec:
                            """ Параметры декларации """
                            for row in range(row_range):
                                """ проверка на наличие данных """
                                row_dec_inp = row_dec_new.copy()
                                row_dec_inp.append(not_employees[0])  # Сотрудник
                                row_dec_inp.append(prop)  # Параметры
                                if (all_dec[count + row][property_dec[prop][0]] is not None) and \
                                        (all_dec[count + row][property_dec[prop][0]] != ''):
                                    for col in property_dec[prop]:
                                        """ Заполнение строки """
                                        if col == 11:
                                            row_dec_inp += (4 * [None])
                                            row_dec_inp.append(all_dec[count + row][col])
                                        elif col == 8:
                                            row_dec_inp += [None]
                                            row_dec_inp.append(all_dec[count + row][col])
                                        else:
                                            row_dec_inp.append(all_dec[count + row][col])

                                    cn += 1  # Следующая предварительная запись
                                    row_dec_inp[0] = cn
                                    # запись в файл
                                    for nr, rd in enumerate(row_dec_inp):
                                        ws_out.cell(row=num_row_out, column=num_col_out + nr).value = rd
                                    num_row_out += 1
                                    # print(row_dec_inp)

                        bar()
                        time.sleep(1)

                        count += row_range  # следующая исходная запись
                        num_row_inp += 1  # следующая главная исходная запись

                    if all_dec[count][1] in not_employees:
                        """ Не сотрудник """
                        not_employees_dec = all_dec[count][1]
                        # поиск записей
                        for row_range in range(len(all_dec[count:])):
                            """ Выделение диапазона строк по данному типу сотрудника """

                            if ((all_dec[count + row_range][1] in not_employees[1:]) and
                                (all_dec[count + row_range][1] != not_employees_dec)) or\
                                    (all_dec[count + row_range][0] == num_row_inp):
                                break
                        if (count + row_range) == (len(all_dec) - 1):
                            row_range += 1
                        row_dec_new = row_dec_new.copy()
                        row_dec_new[0] = cn

                        for prop in property_dec:
                            """ Параметры декларации """
                            for row in range(row_range):
                                """ проверка на наличие данных """
                                row_dec_inp = row_dec_new.copy()
                                row_dec_inp.append(all_dec[count][1])  # Сотрудник
                                row_dec_inp.append(prop)
                                if (all_dec[count + row][property_dec[prop][0]] is not None) and \
                                        (all_dec[count + row][property_dec[prop][0]] != ''):
                                    for col in property_dec[prop]:
                                        """ Заполнение строки """
                                        if col == 11:
                                            row_dec_inp += (4 * [None])
                                            row_dec_inp.append(all_dec[count + row][col])
                                        elif col == 8:
                                            row_dec_inp += [None]
                                            row_dec_inp.append(all_dec[count + row][col])
                                        else:
                                            row_dec_inp.append(all_dec[count + row][col])

                                    cn += 1  # Следующая предварительная запись
                                    row_dec_inp[0] = cn
                                    # запись в файл
                                    for nr, rd in enumerate(row_dec_inp):
                                        ws_out.cell(row=num_row_out, column=num_col_out + nr).value = rd
                                    num_row_out += 1
                                    # print(row_dec_inp)

                        bar()
                        time.sleep(1)

                        count += row_range  # следующая исходная запись

            # Сохраним полный список
            file_xls = os.path.join(BASE_DIR, "output\\declaration.xlsx")
            wb_out.save(file_xls)
            # Закроем исходный список
            wb.close()
    return


def pars_docx():
    """
    Парсинг файлов формата docx
    :return:
    """
    print('Парсинг файлов формата docx')


# Путь к каталогу с исходной информацией
dir_input = os.path.join(BASE_DIR, 'input')
# Файлы с расширением xlsx
files = fnmatch.filter(os.listdir(dir_input), "*.xlsx")
pars_xlsx()
# Файлы с расширением docx
files = fnmatch.filter(os.listdir(dir_input), "*.docx")
pars_docx()
